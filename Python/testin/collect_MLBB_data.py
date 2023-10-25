import openpyxl
import time
from selenium import webdriver
from bs4 import BeautifulSoup
import pyautogui
import keyboard
from selenium_stealth import stealth

def main():
    browser = webdriver.Chrome()
    
    # Value for the URL
    url = "https://liquipedia.net/mobilelegends/MPL/Indonesia/Season_12"
    stats_url = "https://liquipedia.net/mobilelegends/MPL/Indonesia/Season_12/Statistics"
    browser.get(url)
    
    print("Press 'DELETE' to start collecting data.")
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Team's_info"
    
    #modify 1 cell... cell = ws['A4']
    #   or ws['A4'] = 4
    #CELL range can be accessed like:
    #   cell_range = ws['A1':'C2']
    #   column_C = ws['C']
    #   row_10 = ws['10']
    #to iterate through all rows or cols of a file can also do:
    #   for row in ws.values:
    #       for value in row:
    #           print(value)
    while True:
        if keyboard.is_pressed('DELETE'):
            print("Collecting")
            page_source = browser.page_source
            soup = BeautifulSoup(page_source, "html.parser")

            # All teams participating
            table = soup.find("div", class_="table-responsive")
            for row in table.find_all('tr', {"data-toggle-area-content":"8"}):
                cell_data = [cell.get_text() for cell in row.find_all(['th', 'td'])]
                ws.append(cell_data)
            
            print("Getting hero data..")
            browser.get(stats_url)
            time.sleep(5)
            page_source_for_stats = browser.page_source
            soup2 = BeautifulSoup(page_source_for_stats, "html.parser")
            
            table2 = soup2.find("table", class_="wikitable table-striped sortable jquery-tablesorter")
            ws2 = wb_obj.create_sheet(title="Hero_Data")
            wb_obj.active = ws2
            header_data = []
            body_data = []
            first_row = table2.find('tr')
            num_cells_in_first_row = len(first_row.find_all(['th', 'td']))

            for row in table2.find_all('tr'):
                cell_data = [cell.get_text() for cell in row.find_all(['th', 'td'])]
                
                if len(cell_data) == num_cells_in_first_row:
                    header_data.append(cell_data)
                else:
                    body_data.append(cell_data)
                
            for row_data in header_data:
                ws2.append(row_data)
            for row_data in body_data:
                ws2.append(row_data)
                
            print("Data collection complete.")
            break
        elif keyboard.is_pressed('ESC'):
            break
        
        time.sleep(1)
    # Close the browser and quit the WebDriver
    browser.quit()
    print("Saving...")
    wb_obj.save('mpl.xlsx')
    print("Done.")

if __name__ == "__main__":
    main()
